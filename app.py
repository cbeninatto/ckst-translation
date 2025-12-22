import re
import math
import base64
from io import BytesIO
from pathlib import Path

import streamlit as st
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase.pdfmetrics import stringWidth

from openpyxl import load_workbook, Workbook
from openpyxl.utils import range_boundaries

try:
    from openai import OpenAI
except ImportError:
    OpenAI = None


# ---------------- BASIC CONFIG ----------------

st.set_page_config(
    page_title="Carelabel & SKU Label Generator",
    layout="wide",
)

ASSETS_DIR = Path("assets")

BRAND_LOGOS = {
    "Arezzo": ASSETS_DIR / "logo_arezzo.png",
    "Anacapri": ASSETS_DIR / "logo_anacapri.png",
    "Schutz": ASSETS_DIR / "logo_schutz.png",
    "Reserva": ASSETS_DIR / "logo_reserva.png",
}

CARE_ICONS_PATH = ASSETS_DIR / "carelabel_icons.png"


# ---------------- IMAGE HELPERS ----------------

def load_image_base64(path: Path):
    if not path.exists():
        return None
    with open(path, "rb") as f:
        return base64.b64encode(f.read()).decode("utf-8")


BRAND_LOGOS_B64 = {name: load_image_base64(p) for name, p in BRAND_LOGOS.items()}
CARE_ICONS_B64 = load_image_base64(CARE_ICONS_PATH)


# ---------------- TRANSLATION / TEXT (COMPOSITION) ----------------

# Basic EN -> PT material replacements, reused in both freeform and component translation
REPLACEMENTS = [
    (r"polyvinyl chloride\s*\(?\s*pvc\s*\)?", "POLICLORETO DE VINILA (PVC)"),
    (r"\bpvc\b", "POLICLORETO DE VINILA (PVC)"),
    (r"polyurethane", "POLIURETANO (PU)"),
    (r"\bpu\b", "POLIURETANO (PU)"),
    (r"polyester", "POLIÉSTER"),
    (r"polyamide", "POLIAMIDA"),
    (r"nylon", "POLIAMIDA"),
    (r"cotton", "ALGODÃO"),
    (r"filler", "ENCHIMENTO"),
    (r"base fabric", "TECIDO BASE"),
    (r"leather", "COURO"),
    (r"metal", "METAL"),
]


def basic_translate_freeform(text: str) -> str:
    """Apply simple regex-based EN -> PT replacements and uppercase result."""
    if not text:
        return ""
    result = text
    for pattern, repl in REPLACEMENTS:
        result = re.sub(pattern, repl, result, flags=re.IGNORECASE)
    return result.upper()


def parse_components_for_normalization(text: str):
    """Parse compositions like '7%Polyurethane 25%PVC 50%Filler 18%Base Fabric'."""
    if not text:
        return []

    # Normalize spacing: "7%Polyurethane" -> "7% Polyurethane"
    cleaned = text.replace("\n", " ")
    cleaned = re.sub(r"%\s*", "% ", cleaned)

    # Pattern: <number>% <description> (stops before next <number>% or end)
    pattern = re.compile(
        r"(\d+(?:\.\d+)?)\s*%\s*([A-Za-zÀ-ÖØ-öø-ÿ ()/\-]+?)(?=(\d+(?:\.\d+)?\s*%|$))"
    )

    components = []
    for m in pattern.finditer(cleaned):
        pct = float(m.group(1))
        desc = m.group(2).strip()
        components.append((pct, desc))
    return components


def normalize_and_translate_composition(text: str) -> str:
    """
    Normalize composition ignoring Filler/Base Fabric in the percentage base,
    renormalize to 100%, and translate materials to PT.
    """
    if not text:
        return ""

    components = parse_components_for_normalization(text)
    if not components:
        # Fallback: freeform replacement only
        return basic_translate_freeform(text)

    main_components = []
    for pct, desc in components:
        d = desc.lower()
        if "filler" in d or "base fabric" in d or "enchimento" in d or "tecido base" in d:
            # ignore for percentage purposes
            continue
        main_components.append((pct, desc))

    if not main_components:
        # Everything was filler/base fabric → fallback
        return basic_translate_freeform(text)

    total = sum(p for p, _ in main_components)
    if total <= 0:
        return basic_translate_freeform(text)

    # Renormalize percentages to 100 with integer output
    floats = [p * 100.0 / total for p, _ in main_components]
    int_parts = [math.floor(f) for f in floats]
    fracs = [f - i for f, i in zip(floats, int_parts)]

    diff = 100 - sum(int_parts)
    if diff > 0:
        # Give +1 to the 'diff' largest fractional parts
        indices = sorted(range(len(fracs)), key=lambda i: fracs[i], reverse=True)
        for i in indices[:diff]:
            int_parts[i] += 1
    elif diff < 0:
        # Remove 1 from the 'abs(diff)' smallest fractional parts
        indices = sorted(range(len(fracs)), key=lambda i: fracs[i])
        for i in indices[: -diff]:
            int_parts[i] -= 1

    # Build translated string
    parts_pt = []
    for (pct_raw, desc), pct_int in zip(main_components, int_parts):
        material_pt = basic_translate_freeform(desc)
        parts_pt.append(f"{pct_int}% {material_pt}")

    return " ".join(parts_pt)


def translate_composition_to_pt(text: str) -> str:
    """Public function used by the app for EXTERIOR / FORRO."""
    return normalize_and_translate_composition(text)


def build_carelabel_text(exterior_pt: str, forro_pt: str) -> str:
    """Fixed Portuguese body with dynamic EXTERIOR / FORRO."""
    text = f"""IMPORTADO POR BTG PACTUAL
COMMODITIES SERTRADING S.A
CNPJ: 04.626.426/0007-00
DISTRIBUIDO POR:
AZZAS 2154 S.A
CNPJ: 16.590.234/0025-43

FABRICADO NA CHINA
SACAREZZO@AREZZO.COM.BR

PRODUTO DE MATERIAL SINTÉTICO
MATÉRIA-PRIMA
EXTERIOR: {exterior_pt}
FORRO: {forro_pt}

PROIBIDO LAVAR NA ÁGUA / NÃO ALVEJAR /
PROIBIDO USAR SECADOR / NÃO PASSAR
A FERRO / NÃO LAVAR A SECO /
LIMPAR COM PANO SECO"""
    return text


# ---------------- WORD WRAPPING (CARELABEL TEXT) ----------------

def wrap_line(text: str, max_width: float, font_name: str = "Helvetica", font_size: float = 4.0):
    """Wrap a single logical line into multiple lines to fit max_width."""
    if not text:
        return [""]
    words = text.split()
    lines = []
    current = ""

    for w in words:
        if not current:
            current = w
            continue
        candidate = current + " " + w
        if stringWidth(candidate, font_name, font_size) <= max_width:
            current = candidate
        else:
            lines.append(current)
            current = w

    if current:
        lines.append(current)

    return lines


# ---------------- PDF GENERATION: CARELABEL ----------------

def create_carelabel_pdf(brand: str, full_text: str) -> bytes:
    """Create a single-page carelabel PDF with safe top/bottom margins."""

    # Page size (80 x 30 mm carelabel, vertical)
    width = 30 * mm   # width
    height = 80 * mm  # height

    buf = BytesIO()
    c = canvas.Canvas(buf, pagesize=(width, height))

    inner_margin_x = 3 * mm

    # Safe zones (no content near edges – for fold/stitch)
    stitch_margin_mm = 7.0
    safe_top_y = height - stitch_margin_mm * mm
    safe_bottom_y = stitch_margin_mm * mm

    # Bands (inside safe zone)
    top_band_mm = 10.0      # logo band (below safe_top)
    icons_band_mm = 6.0     # icons band (above safe_bottom)

    # ---------- LOGO (TOP, BELOW SAFE MARGIN) ----------
    logo_path = BRAND_LOGOS.get(brand)
    logo_bottom_y_for_text = safe_top_y - top_band_mm * mm  # default if no logo

    if logo_path and logo_path.exists():
        logo_img = ImageReader(str(logo_path))
        iw, ih = logo_img.getSize()

        logo_max_height = (top_band_mm - 2.0) * mm   # inner padding
        logo_max_width = width - 2 * inner_margin_x

        scale = min(logo_max_width / iw, logo_max_height / ih)
        draw_w = iw * scale
        draw_h = ih * scale

        # Reserva logo 1.5x taller (bounded by width)
        if brand == "Reserva":
            draw_w *= 1.5
            draw_h *= 1.5
            if draw_w > logo_max_width:
                factor = logo_max_width / draw_w
                draw_w *= factor
                draw_h *= factor

        # Place logo so its top is slightly below the safe_top_y
        gap_from_safe_top = 1.0 * mm
        y_logo = safe_top_y - gap_from_safe_top - draw_h
        x_logo = (width - draw_w) / 2.0

        c.drawImage(
            logo_img,
            x_logo,
            y_logo,
            width=draw_w,
            height=draw_h,
            preserveAspectRatio=True,
            mask="auto",
        )

        # Text can start a bit below the logo
        text_top_limit = y_logo - 2.0 * mm
    else:
        # If logo missing, text area starts below the safe_top zone
        text_top_limit = logo_bottom_y_for_text

    # ---------- ICONS (BOTTOM, ABOVE SAFE MARGIN) ----------
    icons_max_height = (icons_band_mm - 2.0) * mm   # smaller icons
    icons_max_width = width - 2 * inner_margin_x

    if CARE_ICONS_PATH.exists():
        icons_img = ImageReader(str(CARE_ICONS_PATH))
        iw, ih = icons_img.getSize()
        scale_i = min(icons_max_width / iw, icons_max_height / ih)
        draw_w_i = iw * scale_i
        draw_h_i = ih * scale_i

        # Place icons just above the safe_bottom_y
        gap_from_safe_bottom = 1.0 * mm
        y_icons = safe_bottom_y + gap_from_safe_bottom
        x_icons = (width - draw_w_i) / 2.0

        c.drawImage(
            icons_img,
            x_icons,
            y_icons,
            width=draw_w_i,
            height=draw_h_i,
            preserveAspectRatio=True,
            mask="auto",
        )

        text_bottom_limit = y_icons + draw_h_i + 2.0 * mm
    else:
        # If no icons, keep a band at the bottom anyway
        text_bottom_limit = safe_bottom_y + icons_band_mm * mm

    # ---------- TEXT (MIDDLE, WRAPPED + VERTICALLY CENTERED) ----------
    font_size = 4.0         # smaller text
    leading = 5.0           # line spacing in points
    max_text_width = width - 2 * inner_margin_x

    # Wrap each logical line so it fits in max_text_width
    logical_lines = full_text.splitlines()
    wrapped_lines = []
    for line in logical_lines:
        if not line.strip():
            wrapped_lines.append("")
        else:
            wrapped_lines.extend(
                wrap_line(line, max_text_width, "Helvetica", font_size)
            )

    n_lines = len(wrapped_lines) if wrapped_lines else 1
    text_height = max((n_lines - 1), 0) * leading

    available_top = text_top_limit
    available_bottom = text_bottom_limit
    available_height = available_top - available_bottom

    if available_height <= 0:
        # Degenerate case – just start at top limit
        y_start = available_top
    else:
        if text_height >= available_height:
            # Text does not fit nicely, top-align in available area
            y_start = available_top
        else:
            # Vertically center the text block in [available_bottom, available_top]
            y_start = available_top - (available_height - text_height) / 2.0

    text_obj = c.beginText()
    text_obj.setFont("Helvetica", font_size)
    text_obj.setLeading(leading)
    text_obj.setTextOrigin(inner_margin_x, y_start)

    for line in wrapped_lines:
        # Safety check (should not happen when centered)
        if text_obj.getY() <= text_bottom_limit:
            break
        text_obj.textLine(line)

    c.drawText(text_obj)
    c.showPage()
    c.save()

    pdf_bytes = buf.getvalue()
    buf.close()
    return pdf_bytes


# ---------------- PDF GENERATION: SKU LABELS ----------------

def create_sku_labels_pdf(skus) -> bytes:
    """Multi-page PDF for SKU labels (50 x 10 mm, centered text, no box)."""
    width = 50 * mm
    height = 10 * mm

    buf = BytesIO()
    c = canvas.Canvas(buf, pagesize=(width, height))

    for sku in skus:
        sku = sku.strip()
        if not sku:
            continue

        # No border box; font size increased by 2pt (10 -> 12)
        c.setFont("Helvetica", 12)
        c.drawCentredString(width / 2.0, height / 2.0 - 3, sku)

        c.showPage()

    c.save()
    pdf_bytes = buf.getvalue()
    buf.close()
    return pdf_bytes


# ---------------- HTML PREVIEWS (UI ONLY) ----------------

def carelabel_preview_html(full_text: str, brand: str) -> str:
    """Carelabel preview (approximate on-screen view, without border box)."""
    logo_b64 = BRAND_LOGOS_B64.get(brand)
    icons_b64 = CARE_ICONS_B64

    logo_html = (
        f'<img src="data:image/png;base64,{logo_b64}" '
        f'style="max-width:140px; max-height:90px; margin-bottom:6px;" />'
        if logo_b64
        else ""
    )

    icons_html = (
        f'<img src="data:image/png;base64,{icons_b64}" '
        f'style="width:65%; max-height:40px; margin-top:8px;" />'
        if icons_b64
        else ""
    )

    return f"""
    <div style="
        padding:8px 10px;
        width:260px;
        min-height:520px;
        font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;
        ">
        <div style="text-align:center; margin-bottom:8px;">
            {logo_html}
        </div>
        <div style="font-size:9px; line-height:1.35; white-space:pre-wrap;">
            {full_text}
        </div>
        <div style="margin-top:8px; text-align:center;">
            {icons_html}
        </div>
    </div>
    """


def sku_label_preview_html(sku: str) -> str:
    """Simple horizontal SKU preview (no box)."""
    return f"""
    <div style="
        width:300px;
        height:60px;
        display:flex;
        align-items:center;
        justify-content:center;
        font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;
        font-size:22px;
        letter-spacing:2px;
        margin-bottom:8px;
        ">
        {sku}
    </div>
    """


# ---------------- TECHPACK TRANSLATOR HELPERS ----------------

def get_print_area_bounds(ws):
    """Return (min_row, min_col, max_row, max_col) for the first print area of a worksheet."""
    pa = ws.print_area
    if not pa:
        return None

    # Example: "'C500390022'!$A$1:$K$41" or "A1:K41"
    pa_str = str(pa).split(",")[0]  # first area only
    if "!" in pa_str:
        _, pa_str = pa_str.split("!", 1)
    pa_str = pa_str.replace("$", "")
    min_col, min_row, max_col, max_row = range_boundaries(pa_str)
    return min_row, min_col, max_row, max_col


def extract_print_area_table(ws):
    """Extract cell values from the print area into a 2D list (top-left -> table[0][0])."""
    bounds = get_print_area_bounds(ws)
    if not bounds:
        return None

    min_row, min_col, max_row, max_col = bounds
    table = []
    for r in range(min_row, max_row + 1):
        row = []
        for c in range(min_col, max_col + 1):
            v = ws.cell(row=r, column=c).value
            row.append(v)
        table.append(row)
    return table


def write_table_to_sheet(ws_out, table):
    """Write a 2D table (list of rows) to Worksheet ws_out starting at A1."""
    for r_idx, row in enumerate(table, start=1):
        for c_idx, value in enumerate(row, start=1):
            if value is None or value == "":
                continue
            ws_out.cell(row=r_idx, column=c_idx, value=value)


def get_openai_client(api_key: str | None):
    if OpenAI is None:
        raise RuntimeError(
            "The 'openai' package is not installed. "
            "Install it with 'pip install openai' in your environment."
        )
    if api_key:
        return OpenAI(api_key=api_key)
    return OpenAI()


def translate_table_with_openai(table, client, model_name: str):
    """Translate a rectangular table PT -> EN using the OpenAI Responses API."""
    if not table:
        return []

    # Build plain-text table representation
    rows = []
    for row in table:
        str_cells = []
        for value in row:
            if value is None:
                str_cells.append("")
            else:
                str_cells.append(str(value))
        rows.append("\t".join(str_cells))
    table_str = "\n".join(rows)

    instructions = (
        "You are a professional translator from Brazilian Portuguese to English, "
        "specialized in fashion / handbags techpacks.\n"
        "You will receive the contents of an Excel sheet's PRINT AREA as a rectangular table.\n"
        "Columns are separated by TAB characters (\\t) and rows by newlines (\\n).\n"
        "Translate ONLY the natural language text from Portuguese to English.\n"
        "Keep all article codes, NCMs, dates, sizes, measurements, and numeric values EXACTLY as they are.\n"
        "Do NOT add or remove rows or columns. Do NOT reorder rows or columns.\n"
        "Return ONLY the translated table, in the same format: lines separated by newline, cells separated by TAB.\n"
        "Do not add any explanations or comments."
    )

    response = client.responses.create(
        model=model_name,
        instructions=instructions,
        input=table_str,
    )

    translated_str = response.output_text.strip("\n")
    lines = translated_str.splitlines()

    translated_table = []
    expected_cols = max(len(row) for row in table)

    for line in lines:
        cells = line.split("\t")
        if len(cells) < expected_cols:
            cells += [""] * (expected_cols - len(cells))
        elif len(cells) > expected_cols:
            # Merge extras into last cell
            cells = cells[: expected_cols - 1] + [" ".join(cells[expected_cols - 1 :])]
        translated_table.append(cells)

    # Pad or trim rows to match original row count
    while len(translated_table) < len(table):
        translated_table.append([""] * expected_cols)
    if len(translated_table) > len(table):
        translated_table = translated_table[: len(table)]

    return translated_table


def translate_workbook_to_english(file_bytes: bytes, api_key: str | None, model_name: str):
    """Translate only the print areas of each sheet in a workbook and return a new workbook as bytes."""
    wb_in = load_workbook(BytesIO(file_bytes), data_only=True)
    wb_out = Workbook()
    # Remove default sheet
    default_ws = wb_out.active
    wb_out.remove(default_ws)

    client = get_openai_client(api_key)

    processed_any = False

    for sheet_name in wb_in.sheetnames:
        ws_in = wb_in[sheet_name]
        table = extract_print_area_table(ws_in)
        if table is None:
            # No print area defined -> skip
            continue

        translated_table = translate_table_with_openai(table, client, model_name)
        ws_out_sheet = wb_out.create_sheet(title=sheet_name[:31])  # Excel sheet name limit
        write_table_to_sheet(ws_out_sheet, translated_table)
        processed_any = True

    if not processed_any:
        raise ValueError("No worksheets with a defined print area were found in this workbook.")

    out_buf = BytesIO()
    wb_out.save(out_buf)
    return out_buf.getvalue()


# ---------------- SIDEBAR ----------------

st.sidebar.title("Carelabel Toolkit")

brand = st.sidebar.selectbox("Brand (logo for carelabel)", list(BRAND_LOGOS.keys()))

st.sidebar.markdown("---")
st.sidebar.caption(
    "Carelabel PDF: 80×30 mm (vertical, com margens para costura, sem box).\n"
    "SKU labels PDF: 10×50 mm (horizontal, 1 SKU por página).\n"
    "Techpack translator: lê somente a print area de cada aba."
)


# ---------------- MAIN UI ----------------

st.title("Carelabel, SKU & Techpack Tools")

tab_care, tab_sku, tab_techpack = st.tabs(
    ["Carelabel (80×30 mm)", "SKU labels (10×50 mm)", "Techpack translator (.xlsm → EN)"]
)


# ---- CARELABEL TAB ----
with tab_care:
    col_left, col_right = st.columns([1.1, 1.4])

    with col_left:
        st.subheader("Carelabel – Composição")

        family_code = st.text_input(
            "Product family (para nome do arquivo)",
            value="",
            help="Ex.: C500390016 – todas as cores/SKUs desta família usam a mesma carelabel.",
        )

        st.write("### Composition")
        exterior_en = st.text_input(
            "EXTERIOR",
            value="100% PVC",
            help="English ou Português. Ex.: '75% Polyester, 25% Polyvinyl Chloride (PVC)'",
        )
        forro_en = st.text_input(
            "FORRO / LINING",
            value="100% Polyester",
            help="English ou Português. Ex.: '100% Polyester'",
        )

        already_pt = st.checkbox(
            "Composition already in Portuguese (skip auto-translation)",
            value=False,
        )

        generate_care = st.button("Generate carelabel PDF")

    with col_right:
        st.subheader("Preview & PDF")

        if generate_care:
            # Store family in session for optional use in SKU tab
            st.session_state["family_code"] = family_code.strip()

            if already_pt:
                exterior_pt = exterior_en.strip().upper()
                forro_pt = forro_en.strip().upper()
            else:
                exterior_pt = translate_composition_to_pt(exterior_en)
                forro_pt = translate_composition_to_pt(forro_en)

            full_text = build_carelabel_text(exterior_pt, forro_pt)

            # HTML preview
            st.markdown(
                carelabel_preview_html(full_text, brand),
                unsafe_allow_html=True,
            )

            # PDF
            pdf_bytes = create_carelabel_pdf(brand, full_text)
            pdf_name_base = family_code.strip() or "CARELABEL"
            st.download_button(
                "Download carelabel PDF",
                data=pdf_bytes,
                file_name=f"{pdf_name_base} - CARE LABEL.pdf",
                mime="application/pdf",
            )
        else:
            st.info("Preencha a composição e clique em **Generate carelabel PDF**.")


# ---- SKU LABELS TAB ----
with tab_sku:
    if "sku_count" not in st.session_state:
        st.session_state["sku_count"] = 4  # start with 4 fields

    col_left, col_right = st.columns([1.1, 1.6])

    with col_left:
        st.subheader("SKUs para esta carelabel")

        # Suggest family from carelabel tab, if filled
        default_family = st.session_state.get("family_code", "")
        family_code_sku = st.text_input(
            "Product family (para nome do PDF)",
            value=default_family,
            help="Ex.: C500390016 – usada apenas para nome do PDF.",
        )

        if st.button("Add another SKU field"):
            st.session_state["sku_count"] += 1

        sku_values = []
        for i in range(st.session_state["sku_count"]):
            sku_val = st.text_input(
                f"SKU {i + 1}",
                key=f"sku_{i+1}",
                placeholder="Ex.: C5003900160001",
            )
            if sku_val.strip():
                sku_values.append(sku_val.strip())

        generate_skus = st.button("Generate SKU labels PDF")

    with col_right:
        st.subheader("Preview & PDF")

        if generate_skus:
            if not sku_values:
                st.warning("Informe pelo menos um SKU.")
            else:
                # HTML previews
                for sku in sku_values:
                    st.markdown(sku_label_preview_html(sku), unsafe_allow_html=True)

                # PDF
                sku_pdf = create_sku_labels_pdf(sku_values)
                sku_pdf_name = family_code_sku.strip() or "SKUS"
                st.download_button(
                    "Download SKU labels PDF",
                    data=sku_pdf,
                    file_name=f"{sku_pdf_name} - SKU LABELS.pdf",
                    mime="application/pdf",
                )
        else:
            st.info(
                "Digite os SKUs (vários, se quiser) e clique em "
                "**Generate SKU labels PDF**."
            )


# ---- TECHPACK TRANSLATOR TAB ----
with tab_techpack:
    st.subheader("Techpack translator – .xlsm/.xlsx → English")

    uploaded_file = st.file_uploader(
        "Upload techpack (.xlsm ou .xlsx)",
        type=["xlsm", "xlsx"],
    )

    api_key_input = st.text_input(
        "OpenAI API Key (opcional)",
        type="password",
        help="Se vazio, o app usa a variável de ambiente OPENAI_API_KEY.",
    )

    model_name = st.selectbox(
        "Modelo para tradução",
        ["gpt-4.1-mini", "gpt-4.1", "gpt-5.1"],
        index=0,
    )

    translate_btn = st.button("Translate techpack to English (.xlsx)")

    if translate_btn:
        if uploaded_file is None:
            st.error("Por favor, envie um arquivo .xlsm/.xlsx primeiro.")
        else:
            try:
                file_bytes = uploaded_file.read()
                translated_bytes = translate_workbook_to_english(
                    file_bytes=file_bytes,
                    api_key=api_key_input.strip() or None,
                    model_name=model_name,
                )

                out_name_base = uploaded_file.name.rsplit(".", 1)[0]
                out_name = f"{out_name_base}_EN.xlsx"

                st.success("Techpack traduzido com sucesso (apenas print areas de cada aba).")
                st.download_button(
                    "Download translated techpack (.xlsx)",
                    data=translated_bytes,
                    file_name=out_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            except Exception as e:
                st.error(f"Erro ao traduzir o techpack: {e}")
