import io
import re
from typing import Callable, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.cell_range import CellRange

from .excel_convert import convert_office_bytes
from .openai_translate import OpenAITranslator, TranslationItem
from .text_utils import apply_glossary_hard


def _norm(s: str) -> str:
    s = (s or "").replace("\u00A0", " ").strip().upper()
    s = re.sub(r"\s+", " ", s)
    s = s.rstrip(" .:;,-")
    return s


def _parse_print_area(print_area: str) -> List[CellRange]:
    if not print_area:
        return []
    s = str(print_area).strip()
    if not s:
        return []
    parts = [p.strip() for p in s.split(",") if p.strip()]
    ranges: List[CellRange] = []
    for p in parts:
        if "!" in p:
            p = p.split("!", 1)[1]
        p = p.replace("$", "")
        ranges.append(CellRange(p))
    return ranges


def _cell_in_ranges(r: int, c: int, ranges: List[CellRange]) -> bool:
    for cr in ranges:
        if cr.min_row <= r <= cr.max_row and cr.min_col <= c <= cr.max_col:
            return True
    return False


def _union_bounds(ranges: List[CellRange]) -> Tuple[int, int, int, int]:
    min_row = min(cr.min_row for cr in ranges)
    max_row = max(cr.max_row for cr in ranges)
    min_col = min(cr.min_col for cr in ranges)
    max_col = max(cr.max_col for cr in ranges)
    return min_row, max_row, min_col, max_col


def _unmerge_outside_union(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    try:
        to_unmerge = []
        for mr in list(ws.merged_cells.ranges):
            if not (
                mr.min_row >= min_row
                and mr.max_row <= max_row
                and mr.min_col >= min_col
                and mr.max_col <= max_col
            ):
                to_unmerge.append(str(mr))
        for rng in to_unmerge:
            try:
                ws.unmerge_cells(rng)
            except Exception:
                pass
    except Exception:
        pass


def _crop_sheet_to_union(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    # Delete outside bounds; after this, everything remaining is print area union
    if ws.max_row > max_row:
        ws.delete_rows(max_row + 1, ws.max_row - max_row)
    if min_row > 1:
        ws.delete_rows(1, min_row - 1)

    if ws.max_column > max_col:
        ws.delete_cols(max_col + 1, ws.max_column - max_col)
    if min_col > 1:
        ws.delete_cols(1, min_col - 1)

    try:
        ws.print_title_rows = None
        ws.print_title_cols = None
    except Exception:
        pass

    try:
        last_col = get_column_letter(ws.max_column if ws.max_column >= 1 else 1)
        last_row = ws.max_row if ws.max_row >= 1 else 1
        ws.print_area = f"$A$1:${last_col}${last_row}"
    except Exception:
        pass


def _chunk_list(items: List[TranslationItem], chunk_size: int) -> List[List[TranslationItem]]:
    if chunk_size <= 0:
        return [items]
    return [items[i : i + chunk_size] for i in range(0, len(items), chunk_size)]


def _find_afio_headers(ws, ranges: List[CellRange]) -> List[Tuple[int, int]]:
    """
    Find cells equal to 'AFIO' inside the print area.
    Prefer column G (7).
    Returns list of (row, col).
    """
    found: List[Tuple[int, int]] = []
    found_col7: List[Tuple[int, int]] = []
    for cr in ranges:
        for row in ws.iter_rows(min_row=cr.min_row, max_row=cr.max_row, min_col=cr.min_col, max_col=cr.max_col):
            for cell in row:
                if isinstance(cell.value, str) and _norm(cell.value) == "AFIO":
                    found.append((cell.row, cell.column))
                    if cell.column == 7:
                        found_col7.append((cell.row, cell.column))
    return found_col7 if found_col7 else found


def _under_any_afio_header(r: int, c: int, headers: List[Tuple[int, int]]) -> bool:
    return any(c == hc and r > hr for hr, hc in headers)


def translate_workbook_bytes_openpyxl(
    workbook_bytes: bytes,
    translator: OpenAITranslator,
    source_lang: str,
    target_lang: str,
    glossary: Dict[str, str],
    extra_instructions: str,
    on_progress: Optional[Callable[[str, int, int], None]],
    batch_size: int,
) -> bytes:
    """
    Translate xlsx/xlsm bytes using openpyxl.
    Enforces: translate ONLY inside print area; delete everything outside print area union.
    Special rule: Under AFIO column, if cell == 'NA COR', copy column C from same row (post-translation).
    """
    # Try keep_vba=True first (safe for xlsm); if fails, retry without
    try:
        wb = openpyxl.load_workbook(io.BytesIO(workbook_bytes), keep_vba=True, data_only=False)
    except Exception:
        wb = openpyxl.load_workbook(io.BytesIO(workbook_bytes), keep_vba=False, data_only=False)

    sheets = wb.worksheets
    total_sheets = max(1, len(sheets))
    if on_progress:
        on_progress("pages", 0, total_sheets)

    for s_idx, ws in enumerate(sheets, start=1):
        if on_progress:
            on_progress("pages", s_idx, total_sheets)

        ranges = _parse_print_area(getattr(ws, "print_area", "") or "")
        if not ranges:
            continue

        min_row, max_row, min_col, max_col = _union_bounds(ranges)
        _unmerge_outside_union(ws, min_row, max_row, min_col, max_col)

        afio_headers = _find_afio_headers(ws, ranges)
        na_cor_targets: List[Tuple[int, int]] = []  # (row, afio_col)

        items: List[TranslationItem] = []
        coords: List[Tuple[str, int, int]] = []

        for cr in ranges:
            for row in ws.iter_rows(
                min_row=cr.min_row,
                max_row=cr.max_row,
                min_col=cr.min_col,
                max_col=cr.max_col,
            ):
                for cell in row:
                    v = cell.value
                    if not isinstance(v, str):
                        continue

                    raw = v
                    txt = raw.strip()
                    if not txt:
                        continue

                    # Skip formulas
                    if cell.data_type == "f" or txt.startswith("="):
                        continue

                    # AFIO rule: if NA COR, don't translate; later copy column C
                    if afio_headers and _under_any_afio_header(cell.row, cell.column, afio_headers):
                        if _norm(raw) == "NA COR":
                            na_cor_targets.append((cell.row, cell.column))
                            continue

                    item_id = f"{ws.title}!{cell.coordinate}"
                    items.append(TranslationItem(item_id, raw))
                    coords.append((item_id, cell.row, cell.column))

        batches = _chunk_list(items, batch_size)
        num_batches = max(1, len(batches))
        if on_progress:
            on_progress("batches", 0, num_batches)

        mapping: Dict[str, str] = {}
        for b_idx, batch in enumerate(batches, start=1):
            if on_progress:
                on_progress("batches", b_idx - 1, num_batches)

            mapping.update(
                translator.translate_batch(
                    batch,
                    source_lang=source_lang,
                    target_lang=target_lang,
                    glossary=glossary,
                    extra_instructions=extra_instructions,
                )
            )

            if on_progress:
                on_progress("batches", b_idx, num_batches)

        # Write translations back
        for item_id, r, c in coords:
            new_text = mapping.get(item_id)
            if isinstance(new_text, str):
                ws.cell(row=r, column=c).value = apply_glossary_hard(new_text, glossary)

        # AFIO NA COR -> copy column C (3) value AFTER translation
        for r, afio_col in na_cor_targets:
            src_val = ws.cell(row=r, column=3).value
            ws.cell(row=r, column=afio_col).value = "" if src_val is None else src_val

        # Clear inside union but outside actual print blocks
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                if not _cell_in_ranges(r, c, ranges):
                    cell = ws.cell(row=r, column=c)
                    if cell.value is not None:
                        cell.value = None

        # Delete everything outside print area union and reset print area
        _crop_sheet_to_union(ws, min_row, max_row, min_col, max_col)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def translate_excel_to_xls_bytes(
    excel_bytes: bytes,
    input_ext: str,
    translator: OpenAITranslator,
    source_lang: str = "pt-BR",
    target_lang: str = "en",
    glossary: Optional[Dict[str, str]] = None,
    extra_instructions: str = "",
    on_progress: Optional[Callable[[str, int, int], None]] = None,
    batch_size: int = 25,
) -> bytes:
    """
    Accept .xlsm or .xls input and ALWAYS return .xls output.
    Uses LibreOffice:
      - if input is .xls: convert to .xlsx first (openpyxl can't read .xls)
      - translate via openpyxl
      - convert translated workbook to .xls
    """
    glossary = glossary or {}
    input_ext = input_ext.lower().lstrip(".")

    working_bytes = excel_bytes
    working_ext_for_soffice = input_ext

    if input_ext == "xls":
        # Convert legacy xls -> xlsx so openpyxl can read it
        working_bytes = convert_office_bytes(excel_bytes, "xls", "xlsx")
        working_ext_for_soffice = "xlsx"

    translated_openpyxl_bytes = translate_workbook_bytes_openpyxl(
        workbook_bytes=working_bytes,
        translator=translator,
        source_lang=source_lang,
        target_lang=target_lang,
        glossary=glossary,
        extra_instructions=extra_instructions,
        on_progress=on_progress,
        batch_size=batch_size,
    )

    # Convert translated workbook -> xls
    # (macros are not guaranteed to survive .xls output; that's expected)
    return convert_office_bytes(translated_openpyxl_bytes, working_ext_for_soffice, "xls")
